import base64
import os
from tempfile import TemporaryFile

import openpyxl
from odoo import fields, models
from odoo.exceptions import ValidationError


class Importpartner(models.Model):
    _name = "import.partner"
    _description = "Import Contact"

    file = fields.Binary("File Name", required=True)
    file_name = fields.Char('File Name')

    def import_file(self):
        extension = os.path.splitext(self.file_name)[1].lower()
        if extension not in ['.xlsx', '.xls']:
            raise ValidationError(
                "Invalid file format. You are only allowed to upload .xlsx and .xls file.")
        file_content = base64.decodebytes(self.file)
        excel_fileobj = TemporaryFile('wb+')
        excel_fileobj.write(file_content)
        excel_fileobj.seek(0)
        workbook = openpyxl.load_workbook(excel_fileobj, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        headers = [cell.value for cell in sheet[1]]

        data_list = []
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            row_data = dict(zip(headers, row))
            if 'name' and 'email' in headers:
                data_list.append(row_data)
            else:
                raise ValidationError("Missing 'name' or 'email' in row.")

            existing_partner = self.env['res.partner'].search([('email', '=', row[1])], limit=1)
            partner_data = {
                'name': row[0],
                'email': row[1],
                'phone': row[2],
                'vat': row[3],
                'city': row[4],
                'street': row[5],
                'street2': row[6],
                'zip': row[7],
            }
            if existing_partner:
                existing_partner.write(partner_data)
            else:
                self.env['res.partner'].create(partner_data)
